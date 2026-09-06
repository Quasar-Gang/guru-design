"""XLSX parser: same rules as the CSV parser, applied sheet by sheet."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from packages.importers.document import Document
from packages.importers.parsers.csv_parser import rows_to_document
from packages.importers.ports import RawBlob


class XlsxParser:
    """Parse an .xlsx workbook, reading cell values rather than formulas."""

    def supports(self, fmt: str) -> bool:
        return fmt == "xlsx"

    def parse(self, blob: RawBlob) -> Document:
        if not blob.data:
            return Document()
        workbook = load_workbook(io.BytesIO(blob.data), read_only=True, data_only=True)
        try:
            document = Document()
            for sheet in workbook.worksheets:
                rows = [list(row) for row in sheet.iter_rows(values_only=True)]
                if not rows:
                    continue
                headers = [("" if cell is None else str(cell)) for cell in rows[0]]
                if not any(headers):
                    continue
                document = document.merge(rows_to_document(headers, rows[1:]))
            return document
        finally:
            workbook.close()
